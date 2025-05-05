# playwright_nasdaq.py
import datetime, os, time
from playwright.sync_api import sync_playwright
import openpyxl
from openpyxl.styles import Border, Side, Alignment, Font

def setup_excel_sheet(sheet, headers):
    sheet.append(headers)
    for cell in sheet[1]:
        cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.font = Font(bold=True)

def setup_excel():
    workbook = openpyxl.Workbook()
    sheet1 = workbook.active
    sheet1.title = "52weekhigh"
    setup_excel_sheet(sheet1, ['Stock', 'Price', 'Change', 'Volume', 'Date'])
    return workbook, sheet1

def create_worksheet(workbook, title):
    ws = workbook.create_sheet(title=title)
    setup_excel_sheet(ws, ['Stock', 'Price', 'Change', 'Volume', 'Date'])
    return ws

def process_table(table, worksheet, date):
    rows = table.locator("tr.mhcs-st-row.mhcs-pointer").all()
    for row in rows:
        cols = row.locator("td.mhcs-st-col").all()
        data = [col.inner_text().strip() for col in cols]
        if len(data) >= 4:
            del data[2]  # remove unwanted index
            data.append(date)
            worksheet.append(data[1:])
            for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, min_col=1, max_col=worksheet.max_column):
                for cell in row:
                    cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
                    cell.alignment = Alignment(horizontal="center", vertical="center")

def run_scraper():
    with sync_playwright() as p:
        browser = p.chromium.launch(headless=True)
        page = browser.new_page()
        page.goto("https://stockhouse.com/markets/stocks/nasdaq")

        print("Waiting for page load...")
        page.wait_for_timeout(45000)  # same as time.sleep(45)

        print("Clicking 'More' buttons...")
        more_buttons = page.locator(".mhcs-st-more").all()
        for button in more_buttons:
            button.click()
            time.sleep(2)

        tables = page.locator("table.mhcs-st").all()
        workbook, sheet1 = setup_excel()

        date_elem = page.locator("div.mhcs-headers > div:nth-child(2)").first
        raw_date = date_elem.inner_text().strip().split(",")[0].strip()
        date_str = datetime.datetime.strptime(raw_date, '%b %d').strftime('%d-%b')

        print("Processing 52-Week High Gainers")
        process_table(tables[3], sheet1, date_str)

        print("Processing 52-Week Low Decliners")
        ws1 = create_worksheet(workbook, "52weeklow")
        process_table(tables[4], ws1, date_str)

        print("Processing Net Decliners")
        ws2 = create_worksheet(workbook, "NetDecliner")
        process_table(tables[2], ws2, date_str)

        print("Processing Volume Actives")
        ws3 = create_worksheet(workbook, "Volume_Actives")
        process_table(tables[0], ws3, date_str)

        filename = f"52week_{date_str.replace('-', '')}.xlsx"
        output_path = os.path.join("output", filename)
        os.makedirs("output", exist_ok=True)
        workbook.save(output_path)
        print(f"Saved to {output_path}")
        browser.close()
        return output_path

if __name__ == "__main__":
    run_scraper()
